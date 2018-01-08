
"""

Common internal methods used throughout protocols

"""

# standard libraries

# nonstandard libraries
import openpyxl
from openpyxl.utils import get_column_letter

# homegrown libraries

#--------------------------------------#
#           External Methods           #
#--------------------------------------#

def load_reagent(name):
    pass

def load_sequence(name):
    pass

def load_enzyme(name):
    pass

#--------------------------------------#

def convert_volume(unit,total):

    """ Convert volume units """

    units = _volume_units()

    if unit[1] == 'X':
        unit = (total[0]/float(unit[0]),total[1])
    else:
        assert unit[1] in units.keys(),'units not recognized ({})'.format(unit[1])

    unit_converts = [(unit[0]*units[unit[1]]/v,k) for k,v in units.items()]
    unit_converts = [(k,v) for k,v in unit_converts if k >= 0.1 and k < 100]

    if len(unit_converts) != 1: print 'Something is wrong:',unit_converts

    return unit_converts[0]

#--------------------------------------#

def fill_reaction(total,*args):

    """ Finds how to fill reaction to total """

    units = _volume_units()

    reagent_units = [v[0]*units[v[1]]/units[total[1]] for k,v in args]

    return (total[0] - sum(reagent_units),total[1])

#--------------------------------------#

def request_xlsx(name,fname,sheet=''):

    """ Attempt to pull a value from an xlsx, return as dictionary """

    assert fname.endswith('.xlsx'), "fname extension not operable"
    assert isinstance(sheet,str), "sheet name not string"

    # open database file
    wb = openpyxl.load_workbook(fname)

    if sheet == '':
        sheet = wb.get_sheet_names()[0]
    worksheet = wb.get_sheet_by_name(sheet)

    # extract enzyme information
    value = [i for i in range(2,worksheet.max_row)
            if worksheet['{}{}'.format('A',i)].value.strip(' ') ==
            name.decode('utf-8')]

    # check for matched enzymes
    if len(value) == 0:
        print 'No recognized labels ({})'.format(name)
        database = {}
    elif len(value) == 1:
        database = dict([(k.value,v.value) for k,v
                in zip(tuple(worksheet.rows)[0],tuple(worksheet.rows)[value[0]-1])])
        database = _normalize_dict(database)
    else:
        print 'Multiple matching labels, skipping assignment...'
        database = {}

    # return extracted values
    return database

#--------------------------------------#

def dict_update(base_dict,add_dict):

    """ Appends base dictionary with priority towards existing values """

    # interate through keys
    for k,v in base_dict.items():
        if v == None and k in add_dict.keys():
            base_dict[k] = add_dict[k]

    # return base dictionary
    return base_dict

#--------------------------------------#

def obj_update(obj,*args,**kwargs):

    """ Update object using internal settings """

    if hasattr(obj,'settings'):
        [setattr(obj,k,v) for k,v in obj.settings.items()]
    else:
        print 'No settings attribute detected, skipping update...'

#--------------------------------------#
# TODO: consider faster alternatives (no indexing)
#--------------------------------------#

def comp(seq):

    """ Complement input sequence """

    assert isinstance(seq,str),"submitted sequence is not str"
    seq_map = _seq_map()
    return ''.join(seq_map[seq[i]] for i in xrange(len(seq)))

#--------------------------------------#

def rcomp(seq):

    """ Complement input sequence """

    assert isinstance(seq,str),"submitted sequence is not str"
    seq_map = _seq_map()
    return ''.join(seq_map[seq[len(seq) - i - 1]] for i in xrange(len(seq)))

#--------------------------------------#
#           Internal Methods           #
#--------------------------------------#

def _seq_map():

    # TODO: expand mapping to include degenerative codons

    return {'A':'T','T':'A',
            'C':'G','G':'C',
            'N':'N',' ':' ',
            '|':'|','-':'-'}

#--------------------------------------#

def _volume_units():

    """ Volume units dictionary """

    return {
            'nL':1e-9,
            'uL':1e-6,
            'mL':1e-3,
            'L': 1e0
           }

#--------------------------------------#

def _normalize_dict(my_dict):
    new_dict = {}

    # iterate through dictionary items
    for k,v in my_dict.items():

        # check if float
        try:
            new_dict[str(k).lower()] = float(v)

            continue
        except ValueError:
            pass

        # check if list of strs 
        if v.strip(' ').startswith('[') and v.strip(' ').endswith(']'):
            try:
                new_dict[str(k).lower()] = [float(i) for i in v[1:-1].split(',')]
                continue
            except ValueError:
                new_dict[str(k).lower()] = [str(i).strip(' ') for i in v[1:-1].split(',')]
                continue
        else:
            new_dict[str(k).lower()] = str(v).strip(' ')

    # returns new dictionary
    return new_dict

#--------------------------------------#
#               Testing                #
#--------------------------------------#

if __name__ == "__main__":
    seq = 'AAATTTCCCGGGATCG'
    print seq
    print comp(seq)
    print rcomp(seq)


